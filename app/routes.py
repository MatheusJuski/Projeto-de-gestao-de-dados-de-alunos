from flask import current_app, render_template, redirect, url_for, request, flash, send_file, jsonify, request
from flask_login import login_user, logout_user, current_user, login_required
from app.extensions import login
from app.forms import LoginForm
from app.models import User, users
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import pandas as pd
import re
from cachetools import TTLCache
import threading
from datetime import datetime, timedelta
import random
import requests
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import xlsxwriter
import io
from io import BytesIO
import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from fpdf import FPDF
from calendar import monthrange


dados_cache = []
cache = TTLCache(maxsize=1, ttl=300)


def get_google_sheet():
    """Obtém a planilha do Google Sheets de forma segura"""
    scope = [
        'https://www.googleapis.com/auth/spreadsheets',
        'https://www.googleapis.com/auth/drive'
    ]
    
    try:
        creds = ServiceAccountCredentials.from_json_keyfile_name(
            current_app.config['GOOGLE_SHEETS_CREDENTIALS'], 
            scope
        )
        client = gspread.authorize(creds)
        return client.open_by_key(current_app.config['SPREADSHEET_KEY']).sheet1
    except Exception as e:
        current_app.logger.error(f"Erro ao acessar Google Sheets: {str(e)}")
        raise

def atualizar_cache_threaded(app):
    def atualizar():
        with app.app_context():
            try:
                global dados_cache
                sheet = get_google_sheet()
                data = sheet.get_all_records()
                dados_cache = padronizar_dados(pd.DataFrame(data)).to_dict('records')
                current_app.logger.info("Cache atualizado com sucesso")
            except Exception as e:
                current_app.logger.error(f"Erro ao atualizar cache: {str(e)}")
            finally:
                # Agenda a próxima execução após 5 minutos
                threading.Timer(300.0, lambda: atualizar_cache_threaded(app)).start()

    threading.Thread(target=atualizar).start()  

def formatar_cpf(cpf):
    """Formata o CPF como 000.000.000-00"""
    cpf_numeros = limpar_cpf(cpf)
    if len(cpf_numeros) == 11:
        return f"{cpf_numeros[:3]}.{cpf_numeros[3:6]}.{cpf_numeros[6:9]}-{cpf_numeros[9:]}"
    return cpf  # Retorna como está se não for CPF válido

def limpar_cpf(cpf):
    """Remove formatação do CPF"""
    return re.sub(r'[^0-9]', '', str(cpf)) if cpf else ''
def padronizar_dados(df):
    # Matrícula e CPF (mantém do código anterior)
    if 'Matricula' in df.columns:
        df['Matricula'] = df['Matricula'].astype(str).str.zfill(7)
    
    if 'CPF:' in df.columns:
        df['CPF:'] = df['CPF:'].apply(lambda x: formatar_cpf(x))
    
    # Capitalização de nomes
    campos_para_capitalizar = [
        'Nome completo:',
        'Nome da Mãe:',
        'Nome do Pai:',
        'Naturalidade:',
        'Cidade:',
        'Bairro:',
        'Endereço:',
        'Complemento:',
        'Nome para Contato de Emergência:'
    ]
    
    for campo in campos_para_capitalizar:
        if campo in df.columns:
            df[campo] = df[campo].apply(lambda x: capitalizar_nome(x) if pd.notna(x) else x)
    
    # Links de foto (mantém do código anterior)
    if 'Foto' in df.columns:
        df['Foto'] = df['Foto'].apply(converter_link_google_drive)
    
    
    
    return df

def formatar_agencia(agencia):
    if pd.isna(agencia):
        return ''
    agencia_str = str(agencia).strip()
    # Se tiver "-" ou ".", retorna sem modificar
    if '-' in agencia_str or '.' in agencia_str:
        return agencia_str
    # Limpa tudo que não for número e preenche com zeros à esquerda até 4 dígitos
    numeros = re.sub(r'[^0-9]', '', agencia_str)
    return numeros.zfill(4)

def validar_formato_agencia(agencia):
    if agencia == '':
        return True
    padrao = r'^\d{4}(-\d)?$'
    return bool(re.match(padrao, agencia))

def relatorio_agencias_invalidas(df, coluna='Agência:'):
    invalidos = []
    for val in df[coluna].unique():
        if not validar_formato_agencia(str(val).strip()):
            invalidos.append(val)
    return invalidos

# Função para capitalizar os nomes
def capitalizar_nome(nome):
    if not isinstance(nome, str) or not nome.strip():
        return nome
        
    preposicoes = {'da', 'do', 'das', 'dos', 'e', 'com', 'para', 'a', 'o', 'em', 'por', 'à'}
    palavras = nome.split()
    nome_formatado = []
    
    for i, palavra in enumerate(palavras):
        if i > 0 and palavra.lower() in preposicoes:
            nome_formatado.append(palavra.lower())
        else:
            nome_formatado.append(palavra.capitalize())
    
    return ' '.join(nome_formatado)

def obter_dados_curriculos_filtrados():
    return [aluno for aluno in dados_cache if aluno.get('Currículo Lattes:')]


def format_linkedin_url(url):
    if not url or not isinstance(url, str):
        return ''
    
    url = url.strip().lower()
    
    if re.fullmatch(
        r'^https:\/\/www\.linkedin\.com\/(in|pub|company)\/[a-z0-9\-._]+$', 
        url, 
        flags=re.IGNORECASE
    ):
        return url
    
    if re.fullmatch(
        r'^(www\.)?linkedin\.com\/(in|pub|company)\/[a-z0-9\-._]+$', 
        url, 
        flags=re.IGNORECASE
    ):
        return f'https://{url}'
    
    if re.fullmatch(
        r'^(in|pub|company)\/[a-z0-9\-._]+$', 
        url, 
        flags=re.IGNORECASE
    ):
        return f'https://www.linkedin.com/{url}'
    
    match = re.search(
        r'linkedin\.com\/(in|pub|company)\/([a-z0-9\-._]+)', 
        url, 
        flags=re.IGNORECASE
    )
    if match:
        return f'https://www.linkedin.com/{match.group(1)}/{match.group(2)}'
    

    if re.fullmatch(r'^[a-z0-9\-._]+$', url):
        return f'https://www.linkedin.com/in/{url}'
    

    return ''

def get_cached_data():
    if 'dados' in cache:
        return cache['dados']
    else:
        df = pd.DataFrame(dados_cache)
        cache['dados'] = df
        return df
def get_dados():
    df = padronizar_dados(get_cached_data())
    return df.to_dict('records')

def converter_link_google_drive(link):
    if not isinstance(link, str) or not link:
        return None
    match = re.search(r'/d/([a-zA-Z0-9_-]+)', link)
    if match:
        file_id = match.group(1)
        return f"https://drive.google.com/uc?export=view&id={file_id}"
    elif 'drive.google.com/open?id=' in link:
        file_id = link.split('open?id=')[-1]
        return f"https://drive.google.com/uc?export=view&id={file_id}"
    return link
def calcular_tempo(data_admissao):
    try:
        if pd.isna(data_admissao):
            return "Data não informada"
        hoje = datetime.today()
        diff = hoje - data_admissao
        anos = diff.days // 365
        meses = (diff.days % 365) // 30
        dias = (diff.days % 365) % 30
        return f"{anos} ano(s), {meses} mês(es) e {dias} dia(s)"
    except Exception:
        return "Erro"
    
def identificar_documento(doc):
    doc = str(doc).strip()
    so_numeros = re.sub(r'[^0-9]', '', doc)
    if len(so_numeros) == 11:
        return so_numeros.zfill(11)
    else:
        return ''  # Pode tratar diferente se quiser
def obter_aniversariantes(df, mes, ano, sala_selecionada, modo='mes'):
    aniversariantes = []

    icones_bolo = [
                url_for('static', filename='img/bolo.png'),
                url_for('static', filename='img/bolo2.png'),
            ]
    for _, aluno in df.iterrows():
        data_nascimento = aluno.get('Data de Nascimento')
        if pd.isna(data_nascimento):
            continue

        try:
            nascimento = pd.to_datetime(str(data_nascimento), errors='coerce', dayfirst=True)
            aluno_mes = nascimento.month
            aluno_dia = nascimento.day
            idade = ano - nascimento.year

            # Filtro principal
            incluir = False
            if modo == 'mes' and nascimento.month == mes:
                incluir = True
            elif modo == 'local':
                incluir = True

            # Filtro sala
            if incluir and sala_selecionada:
                if aluno.get('Local de trabalho') != sala_selecionada:
                    incluir = False

            if incluir:
                aniversariantes.append({
                    **aluno,
                    'idade': idade,
                    'dia': aluno_dia,
                    'mes': aluno_mes,
                    'icone_bolo': random.choice(icones_bolo)
                })

        except Exception:
            continue

    # Ordenação
    if modo == 'mes':
        aniversariantes.sort(key=lambda x: x['dia'])
    else:
        aniversariantes.sort(key=lambda x: (x['mes'], x['dia']))

    return aniversariantes


def init_routes(app):
    @app.template_filter('formatar_cpf')
    def formatar_cpf_filter(cpf):
        return formatar_cpf(cpf)

    @app.route('/')
    def index():
        if current_user.is_authenticated:
            return redirect(url_for('inicial'))
        return redirect(url_for('login'))


    @app.route('/login', methods=['GET', 'POST'])
    def login():
        if current_user.is_authenticated:
            return redirect(url_for('inicial'))
        form = LoginForm()
        if form.validate_on_submit():
            user = next((u for u in users.values() if u.username == form.username.data), None)
            if user is None or not user.check_password(form.password.data):
                flash('Usuário ou senha inválidos')
                return redirect(url_for('login'))
            login_user(user, remember=form.remember_me.data)
            return redirect(url_for('inicial'))
        return render_template('login.html', title='Login', form=form)

    @app.route('/logout')
    def logout():
        logout_user()
        return redirect(url_for('login'))

    @app.route('/alunos', methods=['GET'])
    @login_required
    def alunos():
        try:
            
            dados_brutos = get_cached_data()
            df = pd.DataFrame(dados_brutos)
            todos_ids = df['Nome completo:'].dropna().astype(str).tolist()
            df['Agência:'] = df['Agência:'].apply(formatar_agencia)

            df['CPF'] = df['CPF:'].fillna('').astype(str)
            df['CPF:'] = df['CPF:'].str.replace(r'[^0-9]', '', regex=True)
            df['CPF:'] = df['CPF:'].apply(lambda x: x.zfill(11) if x else '')

            df = padronizar_dados(df)
            df = df.iloc[::-1] 

            # Filtros padrão
            local_selecionado = request.args.get('local', '').strip()
            responsavel_selecionado = request.args.get('responsavel', '').strip()
            termo_busca = request.args.get('busca', '').strip()
            status_selecionado = request.args.get('Status')
            foto = request.args.get('Foto', '')


            matricula = request.args.get('Matricula', '').strip()
            cpf = request.args.get('CPF:', '').strip()
            nome = request.args.get('nome', '').strip()

            pagina = request.args.get('pagina', 1, type=int)
            per_page = 8

            dados_filtrados = df

            if matricula:
                dados_filtrados = dados_filtrados[dados_filtrados['Matricula'] == matricula]
            elif cpf:
                cpf_limpo = re.sub(r'[^0-9]', '', cpf)
                dados_filtrados = dados_filtrados[dados_filtrados['CPF:'].str.replace(r'[^0-9]', '', regex=True) == cpf_limpo]
            elif nome:
                dados_filtrados = dados_filtrados[dados_filtrados['Nome completo:'].str.strip() == nome]
            else:
                if local_selecionado:
                    dados_filtrados = dados_filtrados[dados_filtrados['Local de trabalho'] == local_selecionado]
                if responsavel_selecionado:
                    dados_filtrados = dados_filtrados[dados_filtrados['Responsavel pela contratação'] == responsavel_selecionado]
                if termo_busca:
                    termo_normalizado = termo_busca.lower()
                    mask = pd.Series(False, index=dados_filtrados.index)
                    for col in dados_filtrados.columns:
                        if pd.api.types.is_string_dtype(dados_filtrados[col]):
                            mask |= dados_filtrados[col].astype(str).str.lower().str.contains(termo_normalizado, na=False)
                    dados_filtrados = dados_filtrados[mask]
                                
            if status_selecionado and status_selecionado != "todos":
                if status_selecionado == "nao_definido":
                    dados_filtrados = dados_filtrados[
                        dados_filtrados['Status'].isna() | (dados_filtrados['Status'].str.strip() == '')
                    ]
                else:
                    dados_filtrados = dados_filtrados[
                        dados_filtrados['Status'] == status_selecionado
                    ]

            locais_disponiveis = sorted(df['Local de trabalho'].dropna().unique())
            responsavels_disponiveis = sorted(df['Responsavel pela contratação'].dropna().unique())

            total = len(dados_filtrados)
            total_paginas = (total + per_page - 1) // per_page
            start = (pagina - 1) * per_page
            end = start + per_page
            dados_paginados = dados_filtrados.iloc[start:end].to_dict('records')

            ids_filtrados = dados_filtrados['Nome completo:'].tolist()
            

            filtros_ativos = bool(
                termo_busca or local_selecionado or responsavel_selecionado or 
                matricula or cpf or nome or 
                (status_selecionado not in [None, '', 'todos'])
            )

            return render_template(
                    'alunos.html',
                    dados=dados_paginados,
                    termo_busca=termo_busca,
                    ids_filtrados=ids_filtrados,
                    todos_ids=ids_filtrados,
                    locais_disponiveis=locais_disponiveis,
                    local_selecionado=local_selecionado,
                    responsavels_disponiveis=responsavels_disponiveis,
                    responsavel_selecionado=responsavel_selecionado,
                    pagina=pagina,
                    total=total,
                    status_selecionado=status_selecionado,
                    filtros_ativos=filtros_ativos,
                    total_paginas=total_paginas,
                    foto=foto
                )

        except Exception as e:
            flash(f'Erro: {str(e)}')
            return render_template(
                'alunos.html',
                dados=[],
                termo_busca='',
                ids_filtrados=[],
                locais_disponiveis=[],
                local_selecionado='',
                responsavels_disponiveis=[],
                responsavel_selecionado='',
                pagina=1,
                status_selecionado='',
                filtros_ativos=False,
                total_paginas=1
            )
        

    @app.route('/export_alunos', methods=['POST'])
    @login_required
    def export_alunos():
        try:
            data = request.json
            selected_ids = data.get('ids', [])
            formato = data.get('formato', 'xlsx')

            if not selected_ids:
                return jsonify({"error": "Nenhum aluno selecionado"}), 400

            dados_brutos = get_cached_data()
            df = pd.DataFrame(dados_brutos)

            # Filtra os alunos selecionados
            df_filtrado = df[df['Nome completo:'].isin(selected_ids)]

            if df_filtrado.empty:
                return jsonify({"error": "Nenhum aluno encontrado para os nomes fornecidos."}), 404

            # Campos organizados por grupo
            campos = {
                "Dados Pessoais": [
                    "Nome completo:", "Foto", "Data de Nascimento:", "Naturalidade:",
                    "CPF:", "RG:", "Órgão Emissor", "UF:", "País:", "Data de expedição do RG:"
                ],
                "Contato": [
                    "Número do celular", "Número para contato em caso de emergência",
                    "Nome do contato de emergência:", "E-mail para recebimento de informações:"
                ],
                "Endereço": [
                    "Endereço:", "Nº:", "Complemento:", "Bairro: ", "Cidade:", "CEP:"
                ],
                "Dados Bancários": [
                    "Banco: ", "Agência:", "Tipo de conta:", "Número da Conta:"
                ],
                "Dados Acadêmicos": [
                    "Currículo Lattes:", "Matricula", "LinkedIn "
                ],
            }

            colunas_utilizadas = [col for grupo in campos.values() for col in grupo]
            df_filtrado = df_filtrado[colunas_utilizadas]

            if formato == 'xlsx':
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                    df_filtrado.to_excel(writer, index=False, sheet_name='Alunos')
                    worksheet = writer.sheets['Alunos']
                    for i, col in enumerate(df_filtrado.columns):
                        max_len = max(df_filtrado[col].astype(str).map(len).max(), len(col))
                        worksheet.set_column(i, i, max_len + 2)
                output.seek(0)
                return send_file(
                    output,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name='Alunos_selecionados.xlsx'
                )

            elif formato == 'pdf':
                from fpdf import FPDF

                class PDF(FPDF):
                    def aluno_info(self, aluno):
                        self.set_left_margin(15)
                        self.set_right_margin(15)

                        # Grupo: Dados Pessoais
                        self.set_font("Arial", 'B', 11)
                        self.cell(0, 6, "Dados Pessoais", ln=True)
                        self.set_font("Arial", '', 10)
                        for campo in campos["Dados Pessoais"]:
                            valor = aluno.get(campo, "")
                            self.multi_cell(0, 4, f"{campo} {valor}")
                        self.ln(3)

                        # Grupo: Contato
                        self.set_font("Arial", 'B', 11)
                        self.cell(0, 6, "Contato", ln=True)
                        self.set_font("Arial", '', 10)

                        celular = aluno.get("Número do celular", "")
                        email = aluno.get("E-mail para recebimento de informações:", "")
                        tel_emergencia = aluno.get("Número para contato em caso de emergência", "")
                        nome_emergencia = aluno.get("Nome do contato de emergência:", "")

                        self.cell(0, 5, f"Celular: {celular}", ln=True)
                        self.cell(0, 5, f"E-mail: {email}", ln=True)
                        self.cell(0, 5, f"Contato de Emergência: {tel_emergencia} Falar com: {nome_emergencia}", ln=True)
                        self.ln(2)


                        # Grupo: Endereço
                        self.set_font("Arial", 'B', 11)
                        self.cell(0, 6, "Endereço", ln=True)
                        self.set_font("Arial", '', 9)
                        endereco = f"{aluno.get('Endereço:', '')}, Nº {aluno.get('Nº:', '')}, {aluno.get('Complemento:', '')}, "
                        endereco += f"{aluno.get('Bairro: ', '')}, {aluno.get('Cidade:', '')} - CEP {aluno.get('CEP:', '')}"
                        self.multi_cell(0, 4, f"{campo} {valor}")
                        self.ln(3)

                        # Grupo: Dados Bancários
                        self.set_font("Arial", 'B', 11)
                        self.cell(0, 6, "Dados Bancários", ln=True)
                        self.set_font("Arial", '', 9)
                        bancario = f"{aluno.get('Banco: ', '')} | Ag: {aluno.get('Agência:', '')} | "
                        bancario += f"{aluno.get('Tipo de conta:', '')} | Conta: {aluno.get('Número da Conta:', '')}"
                        self.multi_cell(0, 4, f"{campo} {valor}")
                        self.ln(3)

                        # Grupo: Dados Acadêmicos
                        self.set_font("Arial", 'B', 11)
                        self.cell(0, 6, "Dados Acadêmicos", ln=True)
                        self.set_font("Arial", '', 10)
                        for campo in campos["Dados Acadêmicos"]:
                            valor = aluno.get(campo, "")
                            self.multi_cell(0, 4, f"{campo} {valor}")
                        self.ln(4)
        
                pdf = PDF()
                pdf.set_auto_page_break(auto=False, margin=10)
                pdf.set_font("Arial", size=10)

                alunos = df_filtrado.to_dict(orient='records')

                for i, aluno in enumerate(alunos):
                    if i % 2 == 0:
                        pdf.add_page()

                    # Garante que há uma página ativa antes de escrever
                    if pdf.page_no() == 0:
                        pdf.add_page()

                    y_start = 10 if i % 2 == 0 else pdf.h / 2
                    pdf.set_y(y_start)

                    pdf.set_font("Arial", 'B', 13)
                    pdf.cell(0, 8, f"Integrante {i + 1}: {aluno.get('Nome completo:', '')}", ln=True)
                    pdf.set_font("Arial", size=10)
                    pdf.aluno_info(aluno)
                    pdf.ln(2)

                output = io.BytesIO(pdf.output(dest='S').encode('latin1'))
                output.seek(0)

                if len(alunos) == 1:
                    nome_arquivo = f"{alunos[0].get('Nome completo:', 'Aluno')}.pdf"
                else:
                    nome_arquivo = "Informações dos integrantes.pdf"

                return send_file(
                    output,
                    mimetype='application/pdf',
                    as_attachment=True,
                    download_name=nome_arquivo
                )

            else:
                            return jsonify({"error": "Formato inválido. Use 'xlsx' ou 'pdf'."}), 400

        except Exception as e:
            return jsonify({"error": str(e)}), 500



    @app.route('/buscar')
    @login_required
    def buscar():
        try:
            termo_busca = request.args.get('q', '').strip()
            campo_busca = request.args.get('field', 'all').strip()
            local_selecionado = request.args.get('local', '').strip()
            responsavel_selecionado = request.args.get('responsavel', '').strip()

            df = pd.DataFrame(get_cached_data())
            df = padronizar_dados(df)

            df['Agência:'] = df['Agência:'].apply(formatar_agencia)
            df = df.iloc[::-1]  # Mais recentes primeiro

            if local_selecionado:
                df = df[df['Local de trabalho'] == local_selecionado]
            if responsavel_selecionado:
                df = df[df['Responsavel pela contratação'] == responsavel_selecionado]

            if termo_busca:
                termo_normalizado = termo_busca.lower()
                mask = pd.Series(False, index=df.index)

                if campo_busca == 'all':
                    for col in df.columns:
                        if pd.api.types.is_string_dtype(df[col]):
                            if col == 'CPF:':
                                termo_limpo = re.sub(r'[^0-9]', '', termo_busca)
                                mask |= df[col].astype(str).str.replace(r'[^0-9]', '', regex=True).str.contains(termo_limpo, na=False)
                            else:
                                mask |= df[col].astype(str).str.lower().str.contains(termo_normalizado, na=False)
                else:
                    if campo_busca in df.columns:
                        if campo_busca == 'CPF:':
                            termo_limpo = re.sub(r'[^0-9]', '', termo_busca)
                            mask = df[campo_busca].astype(str).str.replace(r'[^0-9]', '', regex=True).str.contains(termo_limpo, na=False)
                        else:
                            mask = df[campo_busca].astype(str).str.lower().str.contains(termo_normalizado, na=False)

                df = df[mask]

                df['CPF:'] = df['CPF:'].astype(str).str.replace(r'[^0-9]', '', regex=True).str.zfill(11)
                df['CPF:'] = df['CPF:'].apply(lambda x: f"{x[:3]}.{x[3:6]}.{x[6:9]}-{x[9:]}")


            return render_template('_cards.html', dados=df.to_dict('records'))

        except Exception as e:
            return f"<p class='no-results'>Erro na busca: {str(e)}</p>"


    @app.route('/aniversarios')
    @login_required
    def aniversarios():
        try:
            modo = request.args.get('modo', 'mes')
            mes = request.args.get('mes', type=int) or datetime.now().month
            ano = request.args.get('ano', type=int) or datetime.now().year
            sala_selecionada = request.args.get('sala', '').strip()

            nomes_meses_pt = {
                1: 'janeiro', 2: 'fevereiro', 3: 'março', 4: 'abril',
                5: 'maio', 6: 'junho', 7: 'julho', 8: 'agosto',
                9: 'setembro', 10: 'outubro', 11: 'novembro', 12: 'dezembro'
            }
            mes_nome = nomes_meses_pt.get(mes, 'Mês desconhecido')

            df = get_cached_data()
            df = padronizar_dados(df)

            aniversariantes = obter_aniversariantes(df, mes, ano, sala_selecionada, modo)

            aniversarios_json = [
                {
                    'nome': aluno.get('Nome Completo', ''),
                    'data_aniversario': f"{ano}-{aluno['mes']:02d}-{aluno['dia']:02d}",
                    'local': aluno.get('Local de trabalho', '')
                }
                for aluno in aniversariantes
            ]

            salas = sorted(df['Local de trabalho'].dropna().unique())

            return render_template(
                'aniversarios.html',
                aniversariantes=aniversariantes,
                mes=mes,
                ano=ano,
                mes_nome=mes_nome,
                nomes_meses_pt=nomes_meses_pt, 
                salas=salas,
                sala_selecionada=sala_selecionada,
                modo=modo,
                aniversarios_json=json.dumps(aniversarios_json)
            )

        except Exception as e:
            flash(f"Erro ao carregar aniversariantes: {e}")
            return render_template(
                'aniversarios.html',
                aniversariantes=[],
                mes=datetime.now().month,
                ano=datetime.now().year,
                mes_nome='Mês desconhecido',
                salas=[],
                sala_selecionada='',
                modo='mes',
                aniversarios_json='[]'
            )
        

    @app.route('/aniversarios/print_calendario')
    @login_required
    def imprimir_calendario_aniversarios():
        try:
            modo = request.args.get('modo', 'mes')  
            mes_param = request.args.get('mes')
            ano = request.args.get('ano', type=int) or datetime.now().year
            sala_selecionada = request.args.get('sala', '').strip()

            df = get_cached_data()
            df = padronizar_dados(df)

            aniversariantes = []

            nomes_meses_pt = {
                1: 'janeiro', 2: 'fevereiro', 3: 'março', 4: 'abril',
                5: 'maio', 6: 'junho', 7: 'julho', 8: 'agosto',
                9: 'setembro', 10: 'outubro', 11: 'novembro', 12: 'dezembro'
            }

            for _, aluno in df.iterrows():
                data_nascimento = aluno.get('Data de Nascimento:')
                if pd.isna(data_nascimento):
                    continue
                nascimento = pd.to_datetime(str(data_nascimento), errors='coerce', dayfirst=True)
                aluno_mes = nascimento.month
                aluno_dia = nascimento.day

                incluir = False

                if mes_param == "todos":
                    incluir = True
                else:
                    mes = int(mes_param)
                    if nascimento.month == mes:
                        incluir = True

                if incluir and sala_selecionada:
                    if aluno.get('Local de trabalho') != sala_selecionada:
                        incluir = False

                if incluir:
                    aniversariantes.append({
                        'Nome completo:': aluno['Nome completo:'],
                        'dia': aluno_dia,
                        'mes': aluno_mes,
                        'idade': datetime.now().year - nascimento.year,
                        'Local de trabalho': aluno.get('Local de trabalho', ''),
                        'icone_bolo': '/static/img/bolo.png',  
                    })

            if mes_param == "todos":
                mes = None
                mes_nome = "Todos os meses"
                primeiro_dia_semana = None
                total_dias_mes = None
            else:
                mes = int(mes_param)
                mes_nome = nomes_meses_pt.get(mes, 'Mês desconhecido')
                primeiro_dia_semana = (datetime(ano, mes, 1).weekday() + 1) % 7
                import calendar
                total_dias_mes = calendar.monthrange(ano, mes)[1]

            return render_template(
                'aniversarios_calendario_print.html',
                aniversariantes=aniversariantes,
                mes=mes,
                ano=ano,
                mes_nome=mes_nome,
                primeiro_dia_semana=primeiro_dia_semana,
                total_dias_mes=total_dias_mes,
                sala_selecionada=sala_selecionada,
                modo=modo,
                nomes_meses_pt=nomes_meses_pt 
            )

        except Exception as e:
            flash(f"Erro ao gerar calendário para impressão: {e}")
            return redirect(url_for('aniversarios'))

    @app.route('/inicial')
    @login_required
    def inicial():
        try:
            df = padronizar_dados(get_cached_data())

            alunos_ativos = df[df['Status'] == 'Ativo']
            total_alunos = len(alunos_ativos)
            mes_atual = datetime.now().month

            # Aniversariantes do mês atual
            aniversariantes = []
            for _, aluno in alunos_ativos.iterrows():
                data_nascimento = aluno.get('Data de Nascimento:')
                if pd.notna(data_nascimento):
                    nascimento = pd.to_datetime(str(data_nascimento), errors='coerce', dayfirst=True)
                    if nascimento.month == mes_atual:
                        aniversariantes.append(aluno)

            # Últimos 5 alunos adicionados (invertido para mostrar o mais recente primeiro)
            ultimos_alunos = df.tail(3).iloc[::-1].to_dict('records')

            # Próximos aniversariantes (próximos 15 dias)
            proximos_aniversariantes = []
            hoje = datetime.now()
            periodo_dias = 15

            for _, aluno in alunos_ativos.iterrows():
                data_nascimento = aluno.get('Data de Nascimento:')
                if pd.isna(data_nascimento):
                    continue
                try:
                    nascimento = pd.to_datetime(str(data_nascimento), errors='coerce', dayfirst=True)
                    if not nascimento:
                        continue
                    
                    aniversario_este_ano = nascimento.replace(year=hoje.year)

                    if aniversario_este_ano < hoje:
                        aniversario_este_ano = aniversario_este_ano.replace(year=hoje.year + 1)

                    dias_ate_aniversario = (aniversario_este_ano - hoje).days

                    if 0 <= dias_ate_aniversario <= periodo_dias:
                        idade = aniversario_este_ano.year - nascimento.year
                        proximos_aniversariantes.append({
                            'Nome completo:': aluno.get('Nome completo:'),
                            'dia': aniversario_este_ano.day,
                            'mes': aniversario_este_ano.month,
                            'idade': idade,
                            'data_aniversario': aniversario_este_ano.strftime('%d/%m'),
                        })
                except Exception:
                    continue
            
            proximos_aniversariantes.sort(key=lambda x: (x['mes'], x['dia']))

            return render_template(
                'inicial.html',
                total_alunos=total_alunos,
                aniversariantes=aniversariantes,
                aniversariantes_do_mes=len(aniversariantes),
                ultimos_alunos=ultimos_alunos,
                mes=mes_atual,
                proximos_aniversariantes=proximos_aniversariantes  # adiciona aqui para o template
            )

        except Exception as e:
            flash(f"Erro ao carregar o dashboard: {str(e)}")
            return render_template(
                'inicial.html',
                total_alunos=0,
                aniversariantes=[],
                aniversariantes_do_mes=0,
                ultimos_alunos=[],
                mes=0,
                proximos_aniversariantes=[]
            )
    @app.route('/curriculos')
    @login_required
    def curriculos():
        dados = get_dados()
        alunos_com_lattes = []

        for a in dados:
            link = str(a.get("Currículo Lattes:", "")).strip()
            if link.startswith("http") and "lattes.cnpq.br" in link:
                alunos_com_lattes.append(a)

        # ORDEM REVERSA
        alunos_com_lattes.reverse()

        # PAGINAÇÃO
        pagina = int(request.args.get('pagina', 1))
        por_pagina = 20
        total = len(alunos_com_lattes)
        inicio = (pagina - 1) * por_pagina
        fim = inicio + por_pagina

        alunos_paginados = alunos_com_lattes[inicio:fim]
        total_paginas = (total + por_pagina - 1) // por_pagina

        return render_template(
            "curriculos.html",
            alunos=alunos_paginados,
            pagina=pagina,
            total_paginas=total_paginas,
            total_curriculos=total
        )

    
    @app.route('/linkedin')
    @login_required
    def linkedin():
        dados = get_dados()
        alunos_com_linkedin = []
        dados = list(reversed(dados))

        # Filtrar alunos com LinkedIn válido
        for a in dados:
            link = a.get("LinkedIn ", "").strip()
            formatted_link = format_linkedin_url(link)
            if formatted_link:
                aluno_com_link = a.copy()
                aluno_com_link['LinkedIn '] = formatted_link
                alunos_com_linkedin.append(aluno_com_link)

        # Paginação
        pagina = int(request.args.get('pagina', 1))
        por_pagina = 20

        total = len(alunos_com_linkedin)
        inicio = (pagina - 1) * por_pagina
        fim = inicio + por_pagina

        alunos_paginados = alunos_com_linkedin[inicio:fim]
        total_paginas = (total + por_pagina - 1) // por_pagina

        return render_template(
            'linkedin.html',
            alunos=alunos_paginados,
            pagina=pagina,
            total_paginas=total_paginas,
            total_linkedin=total
        )


        
        
    @app.route('/tempo_empresa')
    @login_required
    def tempo_empresa():
        df = get_cached_data()

        if 'Data que ingressou no projeto' not in df.columns:
            return "Coluna 'Data que ingressou no projeto' não encontrada na planilha."

        # Converter coluna para datetime
        df['data_admissao'] = pd.to_datetime(
            df['Data que ingressou no projeto'],
            format='%d/%m/%Y',  # ou o formato correto da sua planilha
            dayfirst=True,
            errors='coerce'
        )


        # Coluna com data formatada em string (ou None se inválida)
        def formatar_data(dt):
            if pd.isna(dt):
                return None
            return dt.strftime('%d/%m/%Y')

        df['data_admissao_str'] = df['data_admissao'].apply(formatar_data)

        # Lista de dicionários para passar ao template
        pessoas = []
        for _, row in df.iterrows():
            data_admissao = row['data_admissao']
            
            if pd.notna(data_admissao):
                hoje = datetime.now()
                dias = (hoje - data_admissao).days
                meses = dias // 30
            else:
                dias = -1
                meses = -1

            pessoas.append({
                'nome': row.get('Nome completo:', ''),
                'local': row.get('Local de trabalho', ''),
                'data_admissao': data_admissao,
                'data_admissao_str': row['data_admissao_str'],
                'tempo': calcular_tempo(data_admissao),
                'tempo_dias': dias,
                'tempo_meses': meses,
                'cpf': row.get('CPF:', '')
            })

        return render_template('tempo_empresa.html', pessoas=pessoas)

    @app.route('/exportar_curriculos_excel', methods=['POST'])
    def exportar_curriculos_excel():
        incluir_foto = request.args.get('foto') == '1'
        nomes_selecionados = request.form.get('selecionados')
        nomes_selecionados = json.loads(nomes_selecionados) if nomes_selecionados else []

        dados_filtrados = [
            aluno for aluno in dados_cache
            if aluno.get('Currículo Lattes:') and aluno.get('Nome completo:') in nomes_selecionados
        ]

        # Gerar arquivo Excel
        wb = Workbook()
        ws = wb.active
        headers = ['Nome completo', 'Currículo Lattes']
        if incluir_foto:
            headers.append('Foto')
        ws.append(headers)

        for aluno in dados_filtrados:
            row = [
                aluno.get('Nome completo:', ''),
                aluno.get('Currículo Lattes:', '')
            ]
            if incluir_foto:
                row.append(aluno.get('Foto', '')) 
            ws.append(row)

        # Ajustar automaticamente largura das colunas
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2


        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True, download_name='curriculos.xlsx')


    @app.route('/exportar_curriculos_pdf', methods=['POST'])
    def exportar_curriculos_pdf():
        incluir_foto = request.args.get('foto') == '1'
        nomes_selecionados = request.form.get('selecionados')
        nomes_selecionados = json.loads(nomes_selecionados) if nomes_selecionados else []

        dados_filtrados = [
            aluno for aluno in dados_cache
            if aluno.get('Currículo Lattes:') and aluno.get('Nome completo:') in nomes_selecionados
        ]

        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()
        pdf.set_font("Arial", size=12)

        for aluno in dados_filtrados:
            nome = aluno.get('Nome completo:', 'Sem nome')
            link = aluno.get('Currículo Lattes:', '')
            pdf.cell(200, 10, txt=f"Nome: {nome}", ln=True)
            pdf.cell(200, 10, txt=f"Lattes: {link}", ln=True)
            if incluir_foto:
                foto = aluno.get('Foto', '')
                if foto:
                    pdf.multi_cell(0, 10, txt=f"Foto (link): {foto}")
                else:
                    pdf.cell(200, 10, txt="Foto: Não disponível", ln=True)
            pdf.ln(5)


        pdf_output = pdf.output(dest='S').encode('latin1') 
        output = io.BytesIO(pdf_output)

        return send_file(output, mimetype='application/pdf',
                        as_attachment=True, download_name='curriculos.pdf')
    
    @app.route('/exportar_linkedin_excel', methods=['POST'])
    def exportar_linkedin_excel():
        nomes_selecionados = request.form.get('selecionados')
        nomes_selecionados = json.loads(nomes_selecionados) if nomes_selecionados else []

        dados_filtrados = [
            aluno for aluno in dados_cache
            if aluno.get('LinkedIn ') and aluno.get('Nome completo:') in nomes_selecionados
        ]

        # Gerar arquivo Excel
        wb = Workbook()
        ws = wb.active
        headers = ['Nome completo', 'LinkedIn ']
        ws.append(headers)


        # Ajustar automaticamente largura das colunas
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2


        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True, download_name='Linkedin.xlsx')


    @app.route('/exportar_linkedin_pdf', methods=['POST'])
    def exportar_linkedin_pdf():
        nomes_selecionados = request.form.get('selecionados')
        nomes_selecionados = json.loads(nomes_selecionados) if nomes_selecionados else []

        dados_filtrados = [
            aluno for aluno in dados_cache
            if aluno.get('LinkedIn ') and aluno.get('Nome completo:') in nomes_selecionados
        ]

        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()
        pdf.set_font("Arial", size=12)

        for aluno in dados_filtrados:
            nome = aluno.get('Nome completo:', 'Sem nome')
            link = aluno.get('LinkedIn ', '')
            pdf.cell(200, 10, txt=f"Nome: {nome}", ln=True)
            pdf.cell(200, 10, txt=f"Linkedin: {link}", ln=True)
            pdf.ln(5)


        pdf_output = pdf.output(dest='S').encode('latin1') 
        output = io.BytesIO(pdf_output)

        return send_file(output, mimetype='application/pdf',
                        as_attachment=True, download_name='Linkedin.pdf')
    


